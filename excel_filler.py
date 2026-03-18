"""
Modulo per la compilazione automatica del file Excel
"Controllo Prefabbricazione" a partire dai dati estratti dal PDF DOP.

Utilizza la modifica diretta dell'XML interno al file xlsx (ZIP)
per preservare immagini, formule e formattazione del template originale.

Mappatura celle Excel (basata sul template):
- G2:  Cliente
- D2:  Elemento assiemato (codici posizione)
- D3:  N° commessa
- D4:  Progetto
- G6:  Data collaudo (più recente dal file Excel marcature)
- G10: Data controllo (riga 5 - Esame visivo saldature)
- I14: Data compilazione
"""

import os
import re
import shutil
import zipfile
from datetime import datetime, timedelta

import openpyxl


# Mappatura celle: chiave = nome campo, valore = cella Excel
CELL_MAP = {
    "numero_scheda":          "C1",
    "cliente":                "G2",
    "elemento_assiemato":     "D2",
    "numero_commessa":        "D3",
    "progetto":               "D4",
    "data_collaudo":          "G6",
    "data_controllo":         "G10",
    "data_compilazione":      "I14",
    "responsabile_saldatura": "E10",
}

# Caratteri non validi in XML 1.0 (esclusi tab, newline, carriage return)
_INVALID_XML_CHARS = re.compile(
    r"[^\x09\x0A\x0D\x20-\uD7FF\uE000-\uFFFD\U00010000-\U0010FFFF]"
)


def _sanitize(text: str) -> str:
    """Rimuove caratteri non validi per XML 1.0."""
    return _INVALID_XML_CHARS.sub("", str(text))


def _next_monday_if_weekend(d: datetime) -> datetime:
    """Se d cade di sabato o domenica, avanza al lunedì successivo."""
    if d.weekday() == 5:    # sabato → +2
        return d + timedelta(days=2)
    elif d.weekday() == 6:  # domenica → +1
        return d + timedelta(days=1)
    return d


def _compute_g7_g8_g9(g6_str: str, g10_str: str) -> dict:
    """
    Calcola le date derivate G7, G8, G9 a partire da G6 e G10.
    Restituisce {} se le date non sono valide o l'ordine G6<=G7<G8<G9<G10 non è rispettato.
    """
    fmt = "%d/%m/%Y"
    try:
        g6  = datetime.strptime(g6_str, fmt)
        g10 = datetime.strptime(g10_str, fmt)
    except (ValueError, TypeError):
        return {}
    g7 = g6
    g8 = _next_monday_if_weekend(g6  + timedelta(days=2))
    g9 = _next_monday_if_weekend(g10 - timedelta(days=2))
    if not (g8 > g7 and g9 > g8 and g9 < g10):
        return {}
    return {
        "G7": g7.strftime(fmt),
        "G8": g8.strftime(fmt),
        "G9": g9.strftime(fmt),
    }


def fill_excel(
    template_path: str,
    output_path: str,
    dop_data: dict,
    manual_data: dict = None,
    marcature_excel_path: str = "",
    numero_scheda: int = 0,
    distinta_path: str = "",
) -> str:
    """
    Compila il file Excel template con i dati estratti dal PDF DOP.

    Copia il template e modifica solo le celle necessarie a livello XML,
    preservando immagini, formule e formattazione originali.

    Args:
        template_path: Percorso del file Excel template.
        output_path: Percorso dove salvare il file compilato.
        dop_data: Dizionario con i dati estratti dal PDF.
        manual_data: Dizionario con i campi inseriti manualmente (cliente, numero_commessa, progetto).
        marcature_excel_path: Percorso del file Excel con colonna A=marcature, D=date.
        distinta_path: Percorso del file Excel 'Distinta Spedizione' (colonna A=marcature ammesse).

    Returns:
        Percorso del file salvato.
    """
    cells_to_write = {}

    if numero_scheda > 0:
        cells_to_write[CELL_MAP["numero_scheda"]] = f"{numero_scheda:03d}"

    # Filtra le posizioni in base alla Distinta Spedizione, se fornita
    posizioni = dop_data.get("posizioni", [])
    if distinta_path:
        distinta_set = _get_marcature_from_distinta(distinta_path)
        posizioni = [p for p in posizioni if _normalize_code(p) in distinta_set]

    posizioni_stringa = "-".join(posizioni) if posizioni else ""
    if posizioni_stringa:
        cells_to_write[CELL_MAP["elemento_assiemato"]] = posizioni_stringa
    elif dop_data.get("posizioni_stringa") and not distinta_path:
        cells_to_write[CELL_MAP["elemento_assiemato"]] = dop_data["posizioni_stringa"]

    if dop_data.get("data_ddt"):
        cells_to_write[CELL_MAP["data_controllo"]] = dop_data["data_ddt"]
        cells_to_write[CELL_MAP["data_compilazione"]] = dop_data["data_ddt"]

    # Dati inseriti manualmente dall'utente
    if manual_data:
        for field in ("cliente", "numero_commessa", "progetto", "responsabile_saldatura"):
            val = manual_data.get(field, "").strip()
            if val:
                cells_to_write[CELL_MAP[field]] = val

        resp = manual_data.get("responsabile", "").strip()
        if resp:
            for cell in ("E6", "E7", "E8", "E9", "E11", "E12"):
                cells_to_write[cell] = resp

    # Data più recente dal file Excel marcature → G6
    if marcature_excel_path and posizioni:
        data_collaudo = _get_most_recent_date_from_excel(
            marcature_excel_path, posizioni
        )
        if data_collaudo:
            cells_to_write[CELL_MAP["data_collaudo"]] = data_collaudo

    # Date derivate: G7 = G6, G8 = G6+2gg, G9 = G10-2gg (con salto weekend)
    g6_val  = cells_to_write.get(CELL_MAP["data_collaudo"], "")
    g10_val = cells_to_write.get(CELL_MAP["data_controllo"], "")
    if g6_val and g10_val:
        derived = _compute_g7_g8_g9(g6_val, g10_val)
        cells_to_write.update(derived)

    # Copia il template preservando tutto (immagini, macro, ecc.)
    shutil.copy2(template_path, output_path)

    if cells_to_write:
        _patch_xlsx(output_path, cells_to_write)

    return output_path


def _parse_date_from_string(value: str) -> datetime | None:
    """Estrae la data da stringhe come '1 del 07/04/25' o '2 del 09/04/2025'."""
    match = re.search(r"(\d{2}/\d{2}/\d{2,4})", str(value))
    if not match:
        return None
    date_str = match.group(1)
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def _normalize_code(code: str) -> str:
    """Normalizza un codice: uppercase, rimuove spazi e caratteri non alfanumerici."""
    return re.sub(r"[^A-Z0-9]", "", str(code).upper())


def _get_marcature_from_distinta(distinta_path: str) -> set:
    """
    Legge il file Excel 'Distinta Spedizione' e restituisce l'insieme
    delle marcature presenti nella colonna A (tutti i fogli).
    Supporta sia .xlsx che .xls. Restituisce codici normalizzati.
    """
    if not distinta_path or not os.path.isfile(distinta_path):
        return set()

    result = set()
    ext = os.path.splitext(distinta_path)[1].lower()

    if ext == ".xls":
        # Formato vecchio: usa xlrd
        try:
            import xlrd
            wb = xlrd.open_workbook(distinta_path)
        except Exception:
            return set()
        for sheet in wb.sheets():
            for row_idx in range(sheet.nrows):
                val = sheet.cell_value(row_idx, 0)
                if val:
                    normalized = _normalize_code(str(val))
                    if normalized:
                        result.add(normalized)
    else:
        # Formato nuovo .xlsx: usa openpyxl
        try:
            wb = openpyxl.load_workbook(distinta_path, data_only=True)
        except Exception:
            return set()
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                val = row[0] if row else None
                if val is not None:
                    normalized = _normalize_code(str(val))
                    if normalized:
                        result.add(normalized)

    return result


def _get_most_recent_date_from_excel(excel_path: str, posizioni: list) -> str:
    """
    Legge il file Excel delle marcature e restituisce la data più recente
    tra quelle associate alle posizioni indicate.

    Colonna A: marcatura (es. T2, T4)
    Colonna D: stringa data (es. '1 del 07/04/25')
    """
    if not excel_path or not posizioni or not os.path.isfile(excel_path):
        return ""

    posizioni_set = {str(p).strip().upper() for p in posizioni}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
    except Exception:
        return ""

    dates = []
    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0]).strip().upper() if row[0] is not None else ""
        col_d = row[3] if len(row) > 3 else None

        if col_a in posizioni_set and col_d is not None:
            # col_d può essere già un datetime (se la cella Excel è di tipo data)
            if isinstance(col_d, datetime):
                dates.append(col_d)
            else:
                parsed = _parse_date_from_string(str(col_d))
                if parsed:
                    dates.append(parsed)

    if not dates:
        return ""
    return max(dates).strftime("%d/%m/%Y")


def _col_to_num(col_str):
    """Converte lettera colonna in numero (A=1, B=2, ..., AA=27)."""
    result = 0
    for c in col_str.upper():
        result = result * 26 + (ord(c) - 64)
    return result


def _parse_cell_ref(ref):
    """Splitta riferimento cella in (colonna_str, riga_int)."""
    m = re.match(r"^([A-Z]+)(\d+)$", ref.upper())
    if not m:
        raise ValueError(f"Riferimento cella non valido: {ref!r}")
    return m.group(1), int(m.group(2))


def _resolve_merge(ref, merge_ranges):
    """Se ref e' dentro un range unito, restituisce la cella in alto a sinistra."""
    col_s, row = _parse_cell_ref(ref)
    col = _col_to_num(col_s)
    for mr in merge_ranges:
        tl, br = mr.split(":")
        tl_c, tl_r = _parse_cell_ref(tl)
        br_c, br_r = _parse_cell_ref(br)
        if tl_r <= row <= br_r and _col_to_num(tl_c) <= col <= _col_to_num(br_c):
            return tl
    return ref


def _find_sheet_path(zf):
    """Trova il percorso del primo foglio di lavoro nel file xlsx."""
    for name in sorted(zf.namelist()):
        if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
            return name
    return "xl/worksheets/sheet1.xml"


def _xml_escape(text: str) -> str:
    """Escape minimo per contenuto testo XML."""
    return (text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))


def _patch_xlsx(xlsx_path, cells_to_write):
    """
    Modifica le celle specificate nel file xlsx manipolando direttamente
    i byte XML, senza re-serializzare tramite ElementTree (che altera i
    namespace e corrompe il file).
    """
    tmp = xlsx_path + ".tmp"

    try:
        with zipfile.ZipFile(xlsx_path, "r") as zin:
            sheet_path = _find_sheet_path(zin)
            sheet_bytes = zin.read(sheet_path)

            modified_sheet = _patch_sheet_raw(sheet_bytes, cells_to_write)

            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename == sheet_path:
                        zout.writestr(item, modified_sheet)
                    else:
                        zout.writestr(item, zin.read(item.filename))

        os.replace(tmp, xlsx_path)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise


def _patch_sheet_raw(sheet_bytes: bytes, cells_to_write: dict) -> bytes:
    """
    Modifica celle direttamente nella stringa XML preservando intatti
    tutti i namespace, dichiarazioni e attributi originali.
    """
    xml_str = sheet_bytes.decode("utf-8")

    # Estrai merge ranges dalla stringa XML senza parsare con ET
    merge_ranges = re.findall(r'<mergeCell[^>]+ref="([^"]+)"', xml_str)

    # Risolvi merged cells
    resolved = {}
    for ref, val in cells_to_write.items():
        actual_ref = _resolve_merge(ref, merge_ranges)
        resolved[actual_ref] = _sanitize(val)

    for ref, val in resolved.items():
        xml_str = _write_cell_raw(xml_str, ref, val)

    return xml_str.encode("utf-8")


def _write_cell_raw(xml_str: str, cell_ref: str, value: str) -> str:
    """
    Sostituisce o inserisce il valore di una cella direttamente
    nella stringa XML, preservando attributi di stile (s=) esistenti.
    """
    escaped = _xml_escape(value)
    new_content = f'<is><t>{escaped}</t></is>'

    # Cerca la cella: <c ...r="REF"...>...</c>  oppure  <c ...r="REF".../>
    # Usa lookahead per verificare la presenza di r="REF" nel tag di apertura.
    # IMPORTANTE: usa [^>]*? non-greedy e controlla /> PRIMA di >content</c>
    # per evitare che i tag self-closing mangino le celle successive.
    cell_re = re.compile(
        r'<c\b(?=[^>]*\br="' + re.escape(cell_ref) + r'")([^>]*?)(?:/>|>(.*?)</c>)',
        re.DOTALL,
    )

    match = cell_re.search(xml_str)
    if match:
        attrs = match.group(1).rstrip(" /")  # rimuove il / finale dei tag self-closing
        # Rimuovi tipo esistente (t="...") e attributi formula (cm=, vm=)
        attrs = re.sub(r'\s+t="[^"]*"', "", attrs)
        attrs = re.sub(r'\s+(?:cm|vm)="[^"]*"', "", attrs)
        new_cell = f'<c{attrs} t="inlineStr">{new_content}</c>'
        xml_str = xml_str[: match.start()] + new_cell + xml_str[match.end() :]
        return xml_str

    # Cella non trovata: inseriscila nella riga corrispondente
    col_str, row_num = _parse_cell_ref(cell_ref)
    col_num = _col_to_num(col_str)
    new_cell = f'<c r="{cell_ref}" t="inlineStr">{new_content}</c>'

    row_re = re.compile(
        r'(<row\b(?=[^>]*\br="' + str(row_num) + r'")[^>]*>)(.*?)(</row>)',
        re.DOTALL,
    )
    row_match = row_re.search(xml_str)
    if row_match:
        row_open = row_match.group(1)
        row_body = row_match.group(2)
        row_close = row_match.group(3)

        # Inserisci prima della prima cella con colonna > col_num
        insert_pos = len(row_body)
        for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+)\d+"', row_body):
            ec = _col_to_num(re.match(r"([A-Z]+)", cm.group(1)).group(1))
            if ec > col_num:
                insert_pos = cm.start()
                break

        row_body = row_body[:insert_pos] + new_cell + row_body[insert_pos:]
        new_row = row_open + row_body + row_close
        xml_str = xml_str[: row_match.start()] + new_row + xml_str[row_match.end() :]
    else:
        # Inserisci anche la riga in sheetData
        sd_re = re.compile(r'(<sheetData[^>]*>)', re.DOTALL)
        sd_match = sd_re.search(xml_str)
        if sd_match:
            new_row_el = f'<row r="{row_num}">{new_cell}</row>'
            ins = sd_match.end()
            xml_str = xml_str[:ins] + new_row_el + xml_str[ins:]

    return xml_str


def get_cell_map() -> dict:
    """Restituisce la mappatura celle corrente per visualizzazione nella GUI."""
    return {
        "Cliente (G2)": "cliente",
        "Elemento assiemato (D2)": "posizioni_stringa",
        "N° commessa (D3)": "numero_commessa",
        "Progetto (D4)": "progetto",
        "Data collaudo (G6)": "data_collaudo",
        "Data controllo (G10)": "data_ddt",
        "Data compilazione (I14)": "data_ddt",
    }


if __name__ == "__main__":
    print("=== Mappatura celle Excel ===")
    for label, field in get_cell_map().items():
        print(f"  {label} <- campo PDF: {field}")
